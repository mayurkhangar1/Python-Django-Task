from django.shortcuts import render, redirect
from django.views.generic import ListView
from django.core.paginator import Paginator
from django.http import HttpResponse
from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm
import openpyxl

def add_author(request):
    if request.method == 'POST':
        form = AuthorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('author_list')
    else:
        form = AuthorForm()
    return render(request, 'add_author.html', {'form': form})


def add_book(request):
    if request.method == 'POST':
        form = BookForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('book_list')
    else:
        form = BookForm()
    return render(request, 'add_book.html', {'form': form})


def add_borrow_record(request):
    if request.method == 'POST':
        form = BorrowRecordForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('borrow_list')
    else:
        form = BorrowRecordForm()
    return render(request, 'add_borrow_record.html', {'form': form})


class AuthorListView(ListView):
    model = Author
    template_name = 'author_list.html'
    context_object_name = 'authors'
    paginate_by = 20


def book_list(request):
    book_list = Book.objects.all()
    paginator = Paginator(book_list, 20)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'book_list.html', {'page_obj': page_obj})


def borrow_list(request):
    borrow_list = BorrowRecord.objects.all()
    paginator = Paginator(borrow_list, 20)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'borrow_list.html', {'page_obj': page_obj})


def export_library_data(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Excel_data.xlsx"'

    wb = openpyxl.Workbook()

    author_ws = wb.create_sheet(title='Authors')
    book_ws = wb.create_sheet(title='Books')
    borrow_ws = wb.create_sheet(title='Borrow Records')

    author_ws.append(['ID', 'Name', 'Email', 'Bio'])
    authors = Author.objects.all()
    for author in authors:
        author_ws.append([author.id, author.name, author.email, author.bio])

    book_ws.append(['ID', 'Title', 'Genre', 'Published Date', 'Author'])
    books = Book.objects.all()
    for book in books:
        book_ws.append([book.id, book.title, book.genre, book.published_date, book.author.name])

    borrow_ws.append(['ID', 'User Name', 'Book Title', 'Borrow Date', 'Return Date'])
    borrows = BorrowRecord.objects.all()
    for borrow in borrows:
        borrow_ws.append([borrow.id, borrow.user_name, borrow.book.title, borrow.borrow_date, borrow.return_date])

    wb.save(response)
    return response



